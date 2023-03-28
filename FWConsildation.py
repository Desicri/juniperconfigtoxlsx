#FW Consolidation
import openpyxl
import time
import ipaddress
import re
import logging
from urllib.parse import urlparse

#===============================================================
#region Basic Vars
#Basic Vars
site_name = "  PUT SOMETHING HERE"
filepath = r'C: blah blah blah/'
filename = r"conversionOfRawFWRules " + site_name + ".xlsx"
filenamelog = r"conversionOfRawFWRules " + site_name + ".log"
filenameOut = r"Consolidation of FW Rules " + site_name + ".xlsx"
fileText = filepath + "text"
# Load the workbook and select the worksheet
workbook = openpyxl.load_workbook(filename=filepath + filename)
worksheet = workbook['Parsed-Firewall-Pol']

logging.basicConfig(filename=filepath + filenamelog, level=logging.DEBUG)
dots = ""

firewall_entries_list = []
firewall_unique_set = set()

firewall_policies_dict = {}

#a bool we will use in a function later
application_check = False

#endregion
#===============================================================
#region Functions
#Functions
def Consolidate_Cell_Data(cell_value):
    global application_check
    #make sure we aren't testing blanks
    if cell_value is None:
        return
    
    #if our last cell checked (the last time we ran this function) was "Application", Application check will be "True".
    #Application is always before the port value
    elif application_check == True:
        #we access our dictionary and assign the port key the value in the cell
        firewall_policies_dict[policy_value]['port'] += (cell_value + ", ")

        #we then turn this back off
        application_check = False
    
    #If the cell has source address
    elif "source-address-" in cell_value:
        #because the text has something like "Source-address 192.168.1.1" and we don't need the text we remove it by replacing it with nothing but a comma
        #this should leave jsut the IP or subnet
        cell_value =  cell_value.replace("source-address-", "")
        firewall_policies_dict[policy_value]['source_addresses'] += (cell_value)

    elif "destination-address-" in cell_value:
        #if "destination-address" is there we add a comma
        cell_value =  cell_value.replace("destination-address-", "")
        firewall_policies_dict[policy_value]['destination_addresses'] += (cell_value)

        #this should only have one entry, but just incase there is more than one we want to see it. 
    elif "permit"  in cell_value or "deny"  in cell_value:
        firewall_policies_dict[policy_value]['action'] += (cell_value + ", ") 
    
    #if we see the word application the next value is going to be our port service, so we set a flag for that
    #The next time the script runs it will write the next cell as a port
    elif cell_value == "application":
        application_check = True
    
    #This is just a place to put everything else. We probably won't need it, but we reain it just in case
    else: 
        firewall_policies_dict[policy_value]['etc'] += (cell_value + ",")

def is_website_url(url):
    try:
        result = urlparse(url)
        return all([result.scheme and result.netloc])
    except ValueError:
        return False



def is_valid_ip(address):
    try:
        ipaddress.ip_address(address)
        return True
    except ValueError:
        return False

def is_valid_subnet(subnet):
    try:
        ipaddress.ip_network(subnet)
        return True
    except ValueError:
        return False

def is_valid_ipv6(address):
    try:
        ipaddress.IPv6Address(address)
        return True
    except ValueError:
        return False
    
def is_valid_ipv6_subnet(subnet):
    try:
        ipaddress.IPv6Network(subnet)
        return True
    except ValueError:
        return False

def contains_only_numbers_periods_colons_slashes(string):
    pattern = "^[0-9.:/]*$"
    if re.match(pattern, string):
        return True
    else:
        return False

def isItAnAddress(text):
    if is_valid_ipv6(text) == True:
        #print(text, "is an ipv6 address.")
        return (True)
    
    elif is_valid_subnet(text) == True:
        #print(text, "is an ipv4 subnet.")
        return (True)
    
    elif is_valid_ip(text) == True:
        #print(text, "is an Ipv4 address.")
        return (True)

    elif is_valid_ipv6_subnet(text) == True:
        #print(text, "is an Ipv6 subnet.")
        return (True)
    
    elif contains_only_numbers_periods_colons_slashes(text) == True:
        #print(text, "it is probably an address.")
        return (True)
    
    else:
        logging.debug(str(text) + "is not an address. Investigate!")
        #print ("\t No idea!")
        return (False)
    
def nameToIP(addressesItem):
    IP_Address = ""
    
    #iterate through entire worksheet
    for row in range(2, worksheet.max_row + 1):
        #see waht column 1 has in it far as an address_name
        address_name = worksheet.cell(row=row, column=1).value
        #print ( " ")
        #print ("The address name is", address_name)
        
        #If the values match then return the IP address
        if addressesItem.strip() == address_name.strip():
            IP_Address = worksheet.cell(row=row, column=3).value
            #print ("We found ", addressesItem, "and changed it to", IP_Address)
            #listTemp = IP_Address.split('\n')
            
            #for i in listTemp:
            #    i = i.replace(',', '')
            #    i = i.replace(' ', '')
            #    #isItAnAddress(str(i))
            return IP_Address
    print('\t We did not find', addressesItem)
    logging.debug(str(addressesItem) + "was not found!")
    return (addressesItem.strip())


    

#endregion
#=====================================================================================
#region Consolidating Firewall Rules
#We want to take the mutiple entries in our CSV and consolidate them into one entry

#We need to get a lsit of all the policy names
# We iterate through column I starting at row 2. By adding them to a set we remove duplicates. We end up with the policy names
for row in range(2, worksheet.max_row + 1):
    cell_value = worksheet.cell(row=row, column=4).value
    #We add these to a set so we have a unique list
    firewall_unique_set.add(cell_value)


#Iterate through unique Policy names and make dictionary entries based on the name
#So we will end up with a dictionary full of dictioanries named accoridng to policies.
#We can perform lookups based on the dict name later
for policy_name in firewall_unique_set:
    firewall_policies_dict[policy_name] = {'policy_name' : policy_name, 'source_zone' : "", 'source_addresses' : "", 'action' : "", 'port' : "", 'destination_zone' : "", 'destination_addresses' : "", 'etc' : ''}

#add the dictionary
for dict, val in firewall_policies_dict.items():
    #print(dict + "\n Value: ",  val)
    firewall_entries_list.append(dict)


#Now we go through the excel sheet and grab data and classify it. We will then add it to our "sub dictionary."
# Iterate through column I starting at row 2. We will go through each row and grab the info we want
for row in range(2, worksheet.max_row + 1):
    
    #grab the policy name from the row. We will also use this to access the sub dictionary
    #All the policy names for the entries we want should be the same.
    # This code grabs thw value in row in column D. 
    policy_value = worksheet.cell(row=row, column=4).value

    #Source zone values should be the same for every policy
    #Get the value from the current row Column "B" (column 2). These should always be the same so usign an = operator to jsut replace them is fine
    source_zone_value = worksheet.cell(row=row, column=2).value
   
    #Add this to the dictioanry named after the policy name. They both should have the same name
    firewall_policies_dict[policy_value]['source_zone'] = source_zone_value

    #we just repeat the process to build dictionary the additional entries
    #destination_zone should always be the same, so = is fine. 
    destination_zone_value = worksheet.cell(row=row, column=3).value
    firewall_policies_dict[policy_value]['destination_zone'] = destination_zone_value

    #This is where things get tricky. 
    #we have to get the values of all the parameters and run them through a function to consolidate them
    #However, the data and palcement changes. So we just grab a bunch of data and will sort it out later
    para02_value = worksheet.cell(row=row, column=5).value
    para03_value = worksheet.cell(row=row, column=6).value
    para04_value = worksheet.cell(row=row, column=7).value
    para05_value = worksheet.cell(row=row, column=8).value
    para06_value = worksheet.cell(row=row, column=9).value
    para07_value = worksheet.cell(row=row, column=10).value
    
    #Change all these parameters to list
    parameters = [para02_value,  para03_value, para04_value, para05_value, para06_value, para07_value]
    
    #We set application check to false, this will be important when we run our function in a bit
    application_check = False
    
    #We will not run the parameters through a function
    for para in parameters:
        #filter out some useless repeated data
        if para == "match" or para == "then" or para == "applications":
            continue
        else:
            Consolidate_Cell_Data(para)

#Get an idea of the raw data for debugging
#for dict, val in firewall_policies_dict.items():
#    print(dict + "\n Value: ",  val)

# Create a new workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Add headers to the worksheet
headers = ['policy_name', 'source_zone', 'source_addresses', 'action', 'port', 'destination_zone', 'destination_addresses', 'etc']
worksheet.append(headers)

# Iterate through the dictionary values and write them to the worksheet
for value in firewall_policies_dict.values():
    #filter out denies
    if "deny" in value['action']:
        continue
    #we are going through our dicts and converting them into rows in the excel sheet
    else:
        row = [value['policy_name'], value['source_zone'], value['source_addresses'], value['action'], value['port'], value['destination_zone'], value['destination_addresses'], value['etc']]
        worksheet.append(row)

# Iterate over the rows in the worksheet
for row in worksheet.iter_rows():

    # Iterate over the cells in the row and remove any trailing comma and whitespace. Basically, the last entry won't have a comma
    for cell in row:
        cell.value = str(cell.value.rstrip(", "))  

# Save the workbook
workbook.save(filepath + filenameOut)
#endregion
#=================
#region Address Section #1
#================

#Part 1
#We need to get all the addresses from our conversion sheet and consolidate them too

# Load the workbook we created in the previous step for firewall rules and select the worksheet
workbook = openpyxl.load_workbook(filename=filepath + filename)
print ("Loading " + filepath + filename)
worksheet = workbook['Addresses']

#address_entries_list = []
address_unique_set = set()

address_policies_dict = {}

# Iterate through column I starting at row 2. These are the address names, soem are simple addresses and some are sets. We can sort them later
#What we are doing though is creating a unique set of address names. 
for row in range(2, worksheet.max_row + 1):
    cell_value = worksheet.cell(row=row, column=4).value
    #We add these to a set so we have a unique list
    address_unique_set.add(cell_value)


#Iterate through unique address names and make dictionary entries based on the name
for address_name in address_unique_set:
    address_policies_dict[address_name] = {'address_name' : address_name, 'address-type' : '', 'addresses' : "", 'security-zone' : "", 'action' : ""}



# Iterate through column I starting at row 2. These are the policy names
for row in range(2, worksheet.max_row + 1):
    
    address_set_name_value = worksheet.cell(row=row, column=4).value

    #now we get values from the excel sheet
    #Get the value from the current row Column "G" (column 5). These should always be the same so usign an = operator to jsut replace them is fine
    address_value = worksheet.cell(row=row, column=4).value
    #Add this to the dictioanry named after the address name. They both should have the same name
    address_policies_dict[address_value]['address_name'] = address_value

    #we just repeat the process to build dictionary entries
    security_zone_value = worksheet.cell(row=row, column=2).value
    address_policies_dict[address_value]['security-zone'] = security_zone_value

    #we just repeat the process to build dictionary entries
    address_type = worksheet.cell(row=row, column=3).value
    address_policies_dict[address_value]['address-type'] = address_type

     #we just repeat the process to build dictionary entries
    action_taken = worksheet.cell(row=row, column=1).value
    address_policies_dict[address_value]['action'] = action_taken

    if address_type == "address":
        addresses = worksheet.cell(row=row, column=5).value
        address_policies_dict[address_value]['addresses'] += (addresses + ", ")
    
    if address_type == "address-set":
        addresses = worksheet.cell(row=row, column=6).value
        address_policies_dict[address_value]['addresses'] += (addresses + ", ")


#set the new workbook
workbook = openpyxl.load_workbook(filename=filepath + filenameOut)
new_sheet = workbook.create_sheet("Addresses Parsed")
worksheet = new_sheet

# Add headers to the worksheet
headers = ['address_name', 'address-type', 'addresses', 'security-zone', 'action', 'Possible Set within Set', 'Possible Overflow']
worksheet.append(headers)

# Iterate through the dictionary values and write them to the worksheet
for value in address_policies_dict.values():
    
    row = [value['address_name'], value['address-type'], value['addresses'], value['security-zone'], value['action']]
    worksheet.append(row)



# Iterate over the rows in the worksheet
for row in worksheet.iter_rows():

    # Iterate over the cells in the row
    for cell in row:
        if cell.value is not None:
            cell.value = str(cell.value.rstrip(", "))  

# Save the workbook
workbook.save(filepath + filenameOut)
#endregion
#=====================================================================================

#now we have to take the address sets and work through them to amtch them up with addresses
for row in range(2, worksheet.max_row + 1):

    #just so you know something is happening
    if row % 200 == 0:
        dots += "."
        print ("Working..." + dots)

    #make sure it is an address set
    if worksheet.cell(row=row, column=2).value == "address-set":
        #grab the string of addresses
        address_value = worksheet.cell(row=row, column=3).value
        #turn them into a list divded by commas
        split_list = address_value.split(", ")
        new_list = []
        
        for item in split_list:
            #print ("The item is " + item)
            convertedName = nameToIP(item)
            convertedName = convertedName
            #print ("The converted item is ", convertedName)
            new_list.append(convertedName)
        
        
        combined_string = ''.join(new_list)

        #removing some random commas that somehow got in there
        combined_string = combined_string.replace(",", "")
        combined_string = combined_string.replace(" ", "")
        worksheet.cell(row=row, column=3).value =  combined_string

workbook.save(filepath + filenameOut)

#Because some address sets contian other Address sets rather than URLs or IPs we have to find those andmake additional entries. 
count = 0
#We iterate through the notebook again
for row in range(2, worksheet.max_row + 1):

    #just so you know something is happening
    
    if row % 200 == 0:
        dots += "."
        print ("Working..." + dots)

    #make sure it is an address set
    if worksheet.cell(row=row, column=2).value == "address-set":
        #grab the string of addresses
        address_value = worksheet.cell(row=row, column=3).value
        
        #turn them into a list divded by newlines
        split_list = address_value.split('\n')
        originalMinusAddressSets = []   
        new_list = []

        #debug list
        count += 1
        hasSetWithinSet = False
        overFlowWarning = False
        #print (" ")
        #print (worksheet.cell(row=row, column=1).value)
        #print ("Split list begins. The current palce in the list is ", str(count))
        #print (split_list)
        #print ("the list length is: ", str(len(split_list)))
        #print (" ")

        #print ("The items in the splits list are:")
        for item in split_list:
            
            #check if item is empty
            if item == "":
                continue
            #print ("The item is " + item)
            
            if isItAnAddress(item) == True:
                originalMinusAddressSets.append(item)
            else:
                hasSetWithinSet = True
                convertedName = nameToIP(item)
                #print ("The converted item is ", convertedName)
                new_list.append(convertedName.strip())
        
        
        

        #combine the two lists into strings
        combined_string = '\n'.join(originalMinusAddressSets)
        combined_string2 = '\n'.join(new_list)

        #concatenate the two strings
        combined_string = combined_string + '\n' + combined_string2

        # strip space from combined string
        combined_string = combined_string.strip()
        
        #add it to worksheet
        worksheet.cell(row=row, column=3).value = combined_string

        #mark whether or not we had a set within a set
        if hasSetWithinSet == True:
            worksheet.cell(row=row, column=6).value = "Yes"
        

        #check if we have more than 30000 characters in the cell
        if len(combined_string) > 30000 or len(originalMinusAddressSets) + len(new_list) > 900:
            overFlowWarning = True
            print (" ")
            print ("The following cell " + str(worksheet.cell(row=row, column=1).value) + " has over 30000 characters or 900 lines. Please check the cell for errors.")
            logging.debug(str(worksheet.cell(row=row, column=1).value) + " is too big for a cell!")
            print (" ")
        
        if overFlowWarning == True:
            worksheet.cell(row=row, column=7).value = "Yes"

        
        #press a button to continue
        #input("Press Enter to continue...")
                
workbook.save(filepath + filenameOut)        

        

    
r"""
        with open(fileText, "a") as file:
        # Write the string to the file
            for i in split_list:
                print
                file.write(i)
        #Check if you don't have any letters in the entry. If it has no letters you can skip everything else, because ti is just an IP address. 
        

        contains_letter = False

      

        for item in split_list:
            #If we foudn a letter of the address comes back, we continue our loop and basically leave out this item on the list
            if contains_letter == True:
                continue
            
            #Assuming we have not found a letter of an address we do a letter check
            for char in item:  
                if char.isalpha():
                    contains_letter = True
                    print("WE FOUND letters in", item)
                    break
        
            #if contains_letter == False:
                #print("No letters in", item)
                
        
        
        #removing some random commas that somehow got in there
        string_to_split = string_to_split.replace(",", "")
        print ("The string to split is ", string_to_split)
        
        #turn them into a list divded by newlines
        split_list = string_to_split.split('\n')
        new_list = []
        
        for item in split_list:
            
            if isItAnAddress(item) == False:
                print (str(item) + " is not an IP address")
                ip_value = nameToIP(item)
                print("The entry resolves to ", ip_value)
                new_list.append(ip_value)
            else:
                new_list.append(item)
            #convertedName = nameToIP(item)
            #print ("The converted item is ", convertedName)
            #new_list.append(convertedName)
        
        #combined_string = ''.join(new_list)
        #original_string +=  combined_string

        
        #turn them into a list divded by commas
        #original_string.split(', ')
        #my_set = set(original_string)
        #combined_string = ''.join(my_set)

        #worksheet.cell(row=row, column=3).value = combined_string

#workbook.save(filepath + filenameOut)
"""
