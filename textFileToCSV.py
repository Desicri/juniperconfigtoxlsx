#FW Raw to CSV Policy Parser
import openpyxl
from openpyxl import Workbook
import csv
import os

filepath = r'C:/filepath/'
filename = r"2023-03-14_P-config-set.txt"
filename_csv = 'conversionOfRawFWRules.csv'
filename_xlsx = 'conversionOfRawFWRules.xlsx'
listVar = []

# Define the headers for the CSV file
headers = ['action-taken', 'from-zone', 'to-zone', 'policy-name', 'para1', 'para2', 'para3','para4', 'para5' ]

#format the raw text into a csv usable format
with open(filepath+filename, 'r') as file:
    for line in file:
        if "set security policies from-zone OUTSIDE to-zone" in line:
            grepLine = line
            grepLine = grepLine.replace("set security policies", "set-security-policies")
            grepLine = grepLine.replace("from-zone OUTSIDE", "OUTSIDE")
            grepLine = grepLine.replace("to-zone ", "")
            grepLine = grepLine.replace("policy ", "")
            grepLine = grepLine.replace("destination-address ", "destination-address-")
            grepLine = grepLine.replace("source-address ", "source-address-")
            grepLine = grepLine.replace(" ", ",")
            listVar.append(grepLine)



# Open a file to write as CSV
with open(filepath+'conversionOfRawFWRules.csv', mode='w', newline='') as file:
    # Create a CSV writer object with a comma as delimiter
    writer = csv.writer(file, delimiter=',')
    
    writer.writerow(headers)  
    # Iterate through the list and write each string as a row in the CSV file
    for line in listVar:
        row = line.split(',')  # Split the line into a list of values
        writer.writerow(row)  # Write the list as a row in the CSV file

# Read the CSV file into a pandas DataFrame
filepath_csv = os.path.join(filepath, filename_csv)
filepath_xlsx = os.path.join(filepath, filename_xlsx)

with open(filepath_csv, 'r') as csv_file:
    csv_reader = csv.reader(csv_file)
    data = [row for row in csv_reader]

# Create a new workbook
workbook = Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the data to the worksheet
for row in data:
    worksheet.append(row)

# Save the workbook to XLSX file
workbook.save(filepath_xlsx)

#change sheet name
workbook = openpyxl.load_workbook(filepath_xlsx)
sheet = workbook['Sheet']
sheet.title = 'Parsed-Firewall-Pol'
workbook.save(filepath_xlsx)

#======================
#Get Addresses
#======================

#Make a sheet called "addresses" and make it active
address_sheet = workbook.create_sheet('Addresses')
sheet = workbook['Addresses']
workbook.active = sheet
workbook.save(filepath_xlsx)




# Define the headers for the CSV file
headers2 = ['action-taken', 'security-zone', 'address-type', 'address set name', 'addresses1', 'addresses2', 'para3','para4', 'para5' ]

#make sure you clear this!
listVar = []

#format the raw text into a csv usable format
with open(filepath+filename, 'r') as file:
    for line in file:
        if "set security zones security-zone " in line and "address-book" in line:
            grepLine = line
            grepLine = grepLine.replace("set security zones", "set-security-zones")
            grepLine = grepLine.replace("security-zone ", "")
            grepLine = grepLine.replace("address-book ", "")
            grepLine = grepLine.replace("policy ", "")
            grepLine = grepLine.replace("dns-name ", "")
            grepLine = grepLine.replace(" ", ",")
            listVar.append(grepLine)

# Open a file to write as CSV
with open(filepath+'conversionOfAddresses.csv', mode='w', newline='') as file:
    # Create a CSV writer object with a comma as delimiter
    writer = csv.writer(file, delimiter=',')
    
    writer.writerow(headers2)  
    # Iterate through the list and write each string as a row in the CSV file
    for line in listVar:
        row = line.split(',')  # Split the line into a list of values
        writer.writerow(row)  # Write the list as a row in the CSV file

# Join up paths
filepath_csv = os.path.join(filepath, 'conversionOfAddresses.csv')
filepath_xlsx = os.path.join(filepath, filename_xlsx)

with open(filepath_csv, 'r') as csv_file:
    csv_reader = csv.reader(csv_file)
    data = [row for row in csv_reader]


# Select the active worksheet
worksheet = workbook.active

# Write the data to the worksheet
for row in data:
    worksheet.append(row)

workbook.save(filepath_xlsx)
